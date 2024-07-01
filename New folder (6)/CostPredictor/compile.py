# This is the compilation of code but actual code is written in a modular way
# pip3 install customtkinter pandas openpyxl pillow


import customtkinter
import pandas as pd
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from PIL import Image
import tkinter as tk

# Helper Functions
def get_current_fg_color():
    current_mode = customtkinter.get_appearance_mode()
    return customtkinter.ThemeManager.theme["CTkFrame"]["fg_color"][0 if current_mode == "Light" else 1]

# Frame Creation Functions
def create_home_frame(root):
    home_frame = customtkinter.CTkFrame(root, corner_radius=0, fg_color=get_current_fg_color())
    home_frame.grid_columnconfigure(0, weight=1)
    home_frame.grid_rowconfigure((0, 1, 2, 3, 4), weight=1)

    welcome_label = customtkinter.CTkLabel(home_frame, text="Welcome!", font=customtkinter.CTkFont(size=72),
                                           text_color=("black", "white"))
    welcome_label.grid(row=0, column=0, padx=20, pady=10, sticky="ew")

    home_frame_label = customtkinter.CTkLabel(home_frame, text="This is sample GUI created python customeTkinter just for the demonstation purpose.", text_color=("black", "white"))
    home_frame_label.grid(row=2, column=0, padx=20, pady=10, sticky="ew")

    developed_by_label = customtkinter.CTkLabel(home_frame, text="Developed by Ladybug",
                                                font=customtkinter.CTkFont(size=15), text_color=("black", "white"))
    developed_by_label.grid(row=4, column=0, padx=20, pady=10, sticky="ew")

    return home_frame

def create_frame(root, name):
    frame = customtkinter.CTkFrame(root, corner_radius=0, fg_color=get_current_fg_color())
    frame.grid_columnconfigure(0, weight=1)
    label = customtkinter.CTkLabel(frame, text=name, text_color=("black", "white"))
    label.grid(row=0, column=0, padx=20, pady=10)
    return frame

def create_split_data_frame(root):
    split_data_frame = customtkinter.CTkFrame(root, corner_radius=0, fg_color=get_current_fg_color())
    split_data_frame.grid_columnconfigure(0, weight=1)
    split_data_frame.grid_rowconfigure(6, weight=1)  # Ensure the last row stretches to push buttons to the bottom

    input_file_var = customtkinter.StringVar()
    output_folder_var = customtkinter.StringVar()

    def select_input_file():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            input_file_var.set(file_path)
            input_file_label.configure(text=f"Selected file: {os.path.basename(file_path)}")

    def select_output_path():
        folder_path = filedialog.askdirectory()
        if folder_path:
            output_folder_var.set(folder_path)
            output_folder_label.configure(text=f"Selected folder: {os.path.basename(folder_path)}")

    def create_excel_sheets_from_columns():
        input_file = input_file_var.get()
        output_folder = output_folder_var.get()

        if not input_file or not output_folder:
            messagebox.showerror("Error", "Please select both input file and output folder")
            return

        try:
            # Read the input Excel file
            df = pd.read_excel(input_file)

            # Extract column names
            column_names = df.columns.tolist()

            # Create a new Excel workbook
            output_file = f"{output_folder}/output_workbook_columns.xlsx"
            workbook = Workbook()

            # Remove the default sheet created by Workbook if not needed
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            # Iterate through columns and create sheets
            for col_name in column_names:
                sheet = workbook.create_sheet(title=col_name)
                df_sheet = pd.DataFrame(df[col_name])

                # Convert DataFrame to rows and append to sheet
                for r in dataframe_to_rows(df_sheet, index=False, header=True):
                    sheet.append(r)

            # Save the workbook
            workbook.save(output_file)
            messagebox.showinfo("Success", f"Excel workbook created successfully at: {output_file}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def create_excel_sheets_from_checkboxes():
        output_folder = output_folder_var.get()

        if not output_folder:
            messagebox.showerror("Error", "Please select an output folder")
            return

        try:
            # Create a new Excel workbook
            output_file = f"{output_folder}/output_workbook_checkboxes.xlsx"
            workbook = Workbook()

            # Remove the default sheet created by Workbook if not needed
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            # Iterate through checkboxes and create sheets
            for var, name in zip(checkbox_vars, checkbox_names):
                if var.get() == 1:
                    sheet = workbook.create_sheet(title=name)
                    # Add dummy data for demonstration
                    sheet.append([name, "Sample data"])

            # Save the workbook
            workbook.save(output_file)
            messagebox.showinfo("Success", f"Excel workbook created successfully at: {output_file}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    # Input file selection
    customtkinter.CTkLabel(split_data_frame, text="Select input Excel file:").grid(row=0, column=0, padx=20, pady=(20,5),
                                                                                   sticky="ew")
    input_file_button = customtkinter.CTkButton(split_data_frame, text="Browse...", command=select_input_file,
                                                width=100)
    input_file_button.grid(row=1, column=0, padx=20, pady=5, sticky="w")
    input_file_label = customtkinter.CTkLabel(split_data_frame, text="No file selected",
                                              text_color=("gray50", "gray50"), wraplength=400)
    input_file_label.grid(row=1, column=1, padx=20, pady=5, sticky="w")

    # Output folder selection
    customtkinter.CTkLabel(split_data_frame, text="Select output folder:").grid(row=2, column=0, padx=20, pady=(20, 5),
                                                                                sticky="ew")
    output_folder_button = customtkinter.CTkButton(split_data_frame, text="Browse...", command=select_output_path,
                                                   width=100)
    output_folder_button.grid(row=3, column=0, padx=20, pady=5, sticky="w")
    output_folder_label = customtkinter.CTkLabel(split_data_frame, text="No folder selected",
                                                 text_color=("gray50", "gray50"), wraplength=400)
    output_folder_label.grid(row=3, column=1, padx=20, pady=5, sticky="w")

    # Checkbox frame
    checkbox_frame = customtkinter.CTkFrame(split_data_frame, corner_radius=0)
    checkbox_frame.grid(row=4, column=0, columnspan=2, padx=20, pady=40, sticky="ew")
    checkbox_frame.grid_columnconfigure((0, 1), weight=1)  # Configure two columns

    checkbox_names = [
        "Windows", "Doors", "Floors", "Walls", "Ceilings", "Stairs", "Furniture",
        "Rooms", "Structural columns", "Structural framing", "Wall foundation",
        "Columns foundation (footing)", "Multi-schedule 1 (superstructural)",
        "Multi-schedule 2 (Earthwork)", "Multi-schedule 3 (Complete Schedule)"
    ]
    checkbox_vars = [customtkinter.IntVar() for _ in checkbox_names]

    for i, (name, var) in enumerate(zip(checkbox_names, checkbox_vars)):
        checkbox = customtkinter.CTkCheckBox(checkbox_frame, text=name, variable=var)
        checkbox.grid(row=i // 2, column=i % 2, padx=10, pady=5, sticky="w")  # Two checkboxes per row

    # Generate Excel buttons in a single row at the bottom
    buttons_frame = customtkinter.CTkFrame(split_data_frame, corner_radius=0)
    buttons_frame.grid(row=6, column=0, columnspan=2, padx=20, pady=(5, 20), sticky="sew")
    buttons_frame.grid_columnconfigure((0, 1), weight=1)

    generate_from_columns_button = customtkinter.CTkButton(buttons_frame, text="Generate from Columns",
                                                           command=create_excel_sheets_from_columns, width=150)
    generate_from_columns_button.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

    generate_from_checkboxes_button = customtkinter.CTkButton(buttons_frame, text="Generate from Checkboxes",
                                                              command=create_excel_sheets_from_checkboxes, width=150)
    generate_from_checkboxes_button.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

    return split_data_frame

# Navigation Frame Function
def create_navigation_frame(root, logo_image, toggle_theme, select_frame):
    navigation_frame = customtkinter.CTkFrame(root, corner_radius=0, width=200, fg_color=("#D4D4D4","#323232"))
    navigation_frame.grid(row=0, column=0, sticky="nsew")
    root.grid_rowconfigure(0, weight=1)

    navigation_frame_label = customtkinter.CTkLabel(navigation_frame, text="   Ladybug", image=logo_image,
                                                    compound="left", font=customtkinter.CTkFont(size=15, weight="bold"))
    navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)

    home_button = create_nav_button(navigation_frame, "Home", lambda: select_frame("home"))
    sort_button = create_nav_button(navigation_frame, "Sort", lambda: select_frame("sort"))
    split_data_button = create_nav_button(navigation_frame, "Split Data", lambda: select_frame("split_data"))

    home_button.grid(row=1, column=0, sticky="ew", padx=20, pady=5)
    sort_button.grid(row=2, column=0, sticky="ew", padx=20, pady=5)
    split_data_button.grid(row=3, column=0, sticky="ew", padx=20, pady=5)

    # Add a spacer row with high weight to push other elements up
    navigation_frame.grid_rowconfigure(97, weight=1)

    def quit_app():
        root.quit()

    quit_button = customtkinter.CTkButton(navigation_frame, text="Quit", command=quit_app, fg_color="#BA0B13", hover_color="#94070D")
    quit_button.grid(row=99, column=0, padx=20, pady=(10,30), sticky="ew")

    theme_toggle_switch = customtkinter.CTkSwitch(navigation_frame, text="Dark Mode", command=toggle_theme)
    theme_toggle_switch.grid(row=98, column=0, padx=20, pady=10, sticky="sew")

    return home_button, sort_button, split_data_button

def create_nav_button(parent, text, command):
    button = customtkinter.CTkButton(parent, corner_radius=6, height=40, border_spacing=10, text=text,
                                     fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"), anchor="w", command=command)
    button.grid(sticky="ew")
    return button

# Main Application Code
def setup_images():
    image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "images")
    logo_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "logo.png")), size=(30, 30))
    return logo_image

def select_frame(frame_name):
    for frame in frames.values():
        frame.grid_forget()
    frames[frame_name].grid(row=0, column=1, sticky="nsew")  # Make sure content frames stretch
    for button_name, button in buttons.items():
        button.configure(fg_color=("gray75", "gray25") if button_name == frame_name else "transparent")

def toggle_theme():
    current_mode = customtkinter.get_appearance_mode()
    new_mode = "Dark" if current_mode == "Light" else "Light"
    customtkinter.set_appearance_mode(new_mode)

    # Update the background color of the frames
    new_bg_color = get_current_fg_color()
    for frame in frames.values():
        frame.configure(fg_color=new_bg_color)

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Cost Predictor and Smart Decision App")
    root.geometry("900x650")
    root.resizable(False, False)  # Disable window resizing

    root.grid_columnconfigure(0, weight=0)  # Navigation frame column
    root.grid_columnconfigure(1, weight=1)  # Content frame column
    root.grid_rowconfigure(0, weight=1)  # Ensure the row stretches to fill the space

    logo_image = setup_images()
    home_button, sort_button, split_data_button = create_navigation_frame(root, logo_image, toggle_theme, select_frame)

    frames = {
        "home": create_home_frame(root),
        "sort": create_frame(root, "Sort"),
        "split_data": create_split_data_frame(root),
    }
    buttons = {
        "home": home_button,
        "sort": sort_button,
        "split_data": split_data_button,
    }

    select_frame("home")

    root.mainloop()
